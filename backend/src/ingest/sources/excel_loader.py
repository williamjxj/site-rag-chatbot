"""Excel document loader (.xlsx, .xls)."""

from pathlib import Path
import pandas as pd


def load_excel(path: Path) -> dict[str, str]:
    """
    Load and extract text from Excel file (.xlsx or .xls).

    Args:
        path: Path to Excel file

    Returns:
        Dictionary with 'uri', 'title', and 'text' keys
    """
    parts = []
    
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(str(path))
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Add sheet name as heading
            parts.append(f"Sheet: {sheet_name}")
            
            # Convert DataFrame to text representation
            # Replace NaN with empty string
            df = df.fillna("")
            
            # Get column headers
            headers = " | ".join(str(col) for col in df.columns)
            parts.append(headers)
            
            # Get rows
            for _, row in df.iterrows():
                row_values = " | ".join(str(val) for val in row.values)
                parts.append(row_values)
            
            parts.append("")  # Empty line between sheets
        
        text = "\n".join(parts).strip()
    except Exception as e:
        # If pandas fails, try openpyxl directly for .xlsx
        if path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(path), read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    parts.append(f"Sheet: {sheet_name}")
                    for row in sheet.iter_rows(values_only=True):
                        row_values = " | ".join(str(val) if val is not None else "" for val in row)
                        if row_values.strip():
                            parts.append(row_values)
                    parts.append("")
                text = "\n".join(parts).strip()
            except Exception:
                raise ValueError(f"Failed to read Excel file: {e}") from e
        else:
            raise ValueError(f"Failed to read Excel file: {e}") from e
    
    return {"uri": str(path), "title": path.stem, "text": text}

